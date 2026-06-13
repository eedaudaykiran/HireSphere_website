# ============================================================
# views.py  —  HireHub
# All fixes applied:
#   ✅ Duplicate cache import removed
#   ✅ N+1 in employer_dashboard fixed with annotate()
#   ✅ Duplicate job listing views removed (one per category)
#   ✅ company_profile view uses ModelForm
#   ✅ settings view delegates to SettingsService
#   ✅ reports view delegates to ReportService
#   ✅ transaction.atomic added to apply_job
#   ✅ HttpResponse plain text replaced with messages
#   ✅ _build_company_filter_context DRY violation noted
# ============================================================
from .services import ReportService, SettingsService
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.db import transaction
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.core.cache import cache          # ✅ single import — duplicate removed
from functools import wraps
import datetime
import json
import re

from .forms import (
    RegisterForm, LoginForm, EmployerRegisterForm,
    JobForm, CompanyProfileForm, EmployerSettingsForm,
)
from .models import (
    UserProfile, Job, SavedJob, ApplyJob,
    Application, Interview, Message, CompanyProfile, EmployerSettings,
)



# ===================== ROLE DECORATORS =====================

def candidate_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please login first.")
            return redirect('login')
        try:
            profile = request.user.userprofile
        except UserProfile.DoesNotExist:
            messages.error(request, "Profile not found.")
            return redirect('login')
        if profile.role != 'candidate':
            messages.error(request, "This page is for candidates only.")
            return redirect('employer_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def employer_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please login first.")
            return redirect('login')
        try:
            profile = request.user.userprofile
        except UserProfile.DoesNotExist:
            messages.error(request, "Profile not found.")
            return redirect('login')
        if profile.role != 'employer':
            messages.error(request, "This page is for employers only.")
            return redirect('index')
        return view_func(request, *args, **kwargs)
    return wrapper


# ===================== SHARED HELPER: FILTER COUNTS =====================

def get_filter_counts(base_qs):
    """
    Returns sidebar filter counts for job listing pages.
    Cached per unique queryset for 5 minutes to avoid repeated DB hits.
    """
    cache_key = f"filter_counts_{hash(str(base_qs.query))}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    salary_ranges = ['0-3', '3-6', '6-10', '10-15', '15-20', '20-25', '25-30', '30-35']
    salary_counts = {}
    for r in salary_ranges:
        try:
            low, high = r.split('-')
            salary_counts[r] = base_qs.filter(
                min_salary__gte=int(low),
                max_salary__lte=int(high)
            ).count()
        except (ValueError, AttributeError):
            salary_counts[r] = 0

    category_counts = dict(
        base_qs.values('category').annotate(total=Count('id')).values_list('category', 'total')
    )

    location_list = ['Bangalore', 'Delhi', 'Mumbai', 'Hyderabad', 'Pune', 'Chennai']
    location_counts = {
        loc: base_qs.filter(location__icontains=loc).count()
        for loc in location_list
    }

    company_type_counts = dict(
        base_qs.values('company_type').annotate(total=Count('id')).values_list('company_type', 'total')
    )
    role_counts = dict(
        base_qs.values('role_category').annotate(total=Count('id')).values_list('role_category', 'total')
    )
    duration_counts = dict(
        base_qs.exclude(duration__isnull=True).exclude(duration='')
        .values('duration').annotate(total=Count('id')).values_list('duration', 'total')
    )
    education_counts = dict(
        base_qs.exclude(education__isnull=True).exclude(education='')
        .values('education').annotate(total=Count('id')).values_list('education', 'total')
    )
    posted_by_counts = dict(
        base_qs.exclude(posted_by__isnull=True).exclude(posted_by='')
        .values('posted_by').annotate(total=Count('id')).values_list('posted_by', 'total')
    )
    industry_counts = dict(
        base_qs.exclude(industry__isnull=True).exclude(industry='')
        .values('industry').annotate(total=Count('id')).values_list('industry', 'total')
    )
    company_counts = dict(
        base_qs.values('company').annotate(total=Count('id')).values_list('company', 'total')
    )
    stipend_counts = {
        'unpaid': base_qs.filter(min_salary=0, max_salary=0).count(),
        '0-10':   base_qs.filter(min_salary__gte=0,  max_salary__lte=10).count(),
        '10-20':  base_qs.filter(min_salary__gte=10, max_salary__lte=20).count(),
        '20-30':  base_qs.filter(min_salary__gte=20, max_salary__lte=30).count(),
        '30-50':  base_qs.filter(min_salary__gte=30, max_salary__lte=50).count(),
        '50+':    base_qs.filter(min_salary__gte=50).count(),
    }

    result = {
        'salary_counts':       salary_counts,
        'category_counts':     category_counts,
        'location_counts':     location_counts,
        'company_type_counts': company_type_counts,
        'role_counts':         role_counts,
        'duration_counts':     duration_counts,
        'education_counts':    education_counts,
        'posted_by_counts':    posted_by_counts,
        'industry_counts':     industry_counts,
        'company_counts':      company_counts,
        'stipend_counts':      stipend_counts,
    }
    cache.set(cache_key, result, timeout=300)
    return result


# ===================== SHARED HELPER: APPLY ALL FILTERS =====================

def apply_all_filters(qs, request):
    """
    Applies all sidebar filters from GET params to a queryset.
    Returns (filtered_qs, selected_dict) — selected dict is passed to template
    so checkboxes stay ticked after filtering.
    """
    work_modes = request.GET.getlist('work_mode')
    if work_modes:
        qs = qs.filter(work_mode__in=work_modes)

    categories = request.GET.getlist('category')
    if categories:
        qs = qs.filter(category__in=categories)

    locations = request.GET.getlist('location')
    if locations:
        q = Q()
        for loc in locations:
            q |= Q(location__icontains=loc)
        qs = qs.filter(q)

    salaries = request.GET.getlist('salary')
    if salaries:
        q = Q()
        for s in salaries:
            try:
                low, high = s.split('-')
                q |= Q(min_salary__gte=int(low), max_salary__lte=int(high))
            except (ValueError, AttributeError):
                continue
        qs = qs.filter(q)

    experience = request.GET.get('experience')
    if experience and experience != '30':
        try:
            exp_int = int(experience)
            qs = qs.filter(min_experience__lte=exp_int, max_experience__gte=exp_int)
        except ValueError:
            pass

    freshness = request.GET.get('freshness')
    if freshness:
        try:
            cutoff = timezone.now() - datetime.timedelta(days=int(freshness))
            qs = qs.filter(created_at__gte=cutoff)
        except (ValueError, TypeError):
            pass

    company_types = request.GET.getlist('company_type')
    if company_types:
        qs = qs.filter(company_type__in=company_types)

    durations = request.GET.getlist('duration')
    if durations:
        qs = qs.filter(duration__in=durations)

    educations = request.GET.getlist('education')
    if educations:
        qs = qs.filter(education__in=educations)

    posted_by = request.GET.getlist('posted_by')
    if posted_by:
        qs = qs.filter(posted_by__in=posted_by)

    industries = request.GET.getlist('industry')
    if industries:
        qs = qs.filter(industry__in=industries)

    companies = request.GET.getlist('company')
    if companies:
        qs = qs.filter(company__in=companies)

    roles = request.GET.getlist('role_category')
    if roles:
        qs = qs.filter(role_category__in=roles)

    stipends = request.GET.getlist('stipend')
    if stipends:
        q = Q()
        for s in stipends:
            if s == 'unpaid':
                q |= Q(min_salary=0, max_salary=0)
            elif s == '0-10':
                q |= Q(min_salary__gte=0,  max_salary__lte=10)
            elif s == '10-20':
                q |= Q(min_salary__gte=10, max_salary__lte=20)
            elif s == '20-30':
                q |= Q(min_salary__gte=20, max_salary__lte=30)
            elif s == '30-50':
                q |= Q(min_salary__gte=30, max_salary__lte=50)
            elif s == '50+':
                q |= Q(min_salary__gte=50)
        qs = qs.filter(q)

    selected = {
        'selected_work_modes':    work_modes,
        'selected_categories':    categories,
        'selected_locations':     locations,
        'selected_salaries':      salaries,
        'selected_experience':    experience,
        'selected_freshness':     freshness,
        'selected_company_types': company_types,
        'selected_durations':     durations,
        'selected_educations':    educations,
        'selected_posted':        posted_by,
        'selected_industries':    industries,
        'selected_companies':     companies,
        'selected_roles':         roles,
        'selected_stipends':      stipends,
    }
    return qs, selected


# ===================== SHARED HELPER: FULL JOB PAGE =====================

def job_list_view(request, base_qs, template_name):
    """
    Single reusable view used by all 30+ job listing pages.
    Handles filtering, pagination, counts in one place.
    """
    all_jobs = base_qs
    jobs, selected = apply_all_filters(base_qs, request)
    counts = get_filter_counts(all_jobs)

    paginator = Paginator(jobs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'jobs':     page_obj,
        'page_obj': page_obj,
        **selected,
        **counts,
    }
    return render(request, template_name, context)


# ===================== COMPANY PAGE HELPER =====================

def _build_company_filter_context(request, base_qs):
    """
    Builds filter context for company listing pages.
    Uses apply_all_filters for shared filter logic.
    """
    from django.db.models import Min, Max

    # ✅ Reuse apply_all_filters instead of duplicating logic
    jobs, selected = apply_all_filters(base_qs, request)

    companies = (
        jobs
        .values('company', 'company_type', 'location', 'industry', 'nature_of_business')
        .annotate(job_count=Count('id'), min_sal=Min('min_salary'), max_sal=Max('max_salary'))
        .order_by('-job_count')
    )

    profile_map = {
        cp.employer.username.lower(): cp
        for cp in CompanyProfile.objects.all()
    }

    enriched = []
    for c in companies:
        name    = c['company']
        profile = profile_map.get((name or '').lower())
        enriched.append({
            'name':               name,
            'company_type':       c['company_type'],
            'location':           c['location'],
            'industry':           c['industry'],
            'nature_of_business': c['nature_of_business'],
            'job_count':          c['job_count'],
            'logo':               profile.logo if profile else None,
            'description':        profile.description if profile else '',
        })

    all_q                = base_qs
    location_list        = ['Bangalore', 'Delhi', 'Mumbai', 'Hyderabad', 'Pune', 'Chennai']
    location_counts      = {loc: all_q.filter(location__icontains=loc).count() for loc in location_list}
    company_type_counts  = dict(all_q.values('company_type').annotate(total=Count('id')).values_list('company_type', 'total'))
    industry_counts      = dict(all_q.exclude(industry='').values('industry').annotate(total=Count('id')).values_list('industry', 'total'))
    department_counts    = dict(all_q.exclude(department__isnull=True).exclude(department='').values('department').annotate(total=Count('id')).values_list('department', 'total'))
    nob_counts           = dict(all_q.exclude(nature_of_business__isnull=True).values('nature_of_business').annotate(total=Count('id')).values_list('nature_of_business', 'total'))

    salary_ranges  = ['0-3', '3-6', '6-10', '10-15', '15-20', '20-25', '25-30', '30-35']
    salary_counts  = {}
    for r in salary_ranges:
        try:
            lo, hi = r.split('-')
            salary_counts[r] = all_q.filter(min_salary__gte=int(lo), max_salary__lte=int(hi)).count()
        except Exception:
            salary_counts[r] = 0

    context = {
        'companies':              enriched,
        'total_companies':        len(enriched),
        'location_counts':        location_counts,
        'company_type_counts':    company_type_counts,
        'industry_counts':        industry_counts,
        'department_counts':      department_counts,
        'nob_counts':             nob_counts,
        'salary_counts':          salary_counts,
        **selected,                # ✅ unpacked from apply_all_filters
    }
    return jobs, context


# ===================== BASIC VIEWS =====================

def index(request):
    jobs      = Job.objects.filter(is_active=True).order_by('-created_at')
    paginator = Paginator(jobs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/index.html', {'jobs': page_obj})


def search_jobs_page(request):
    query      = request.GET.get('q', '').strip()
    location   = request.GET.get('location', '').strip()
    experience = request.GET.get('experience', '').strip()

    jobs = Job.objects.filter(is_active=True)

    if query:
        jobs = jobs.filter(
            Q(title__icontains=query) |
            Q(company__icontains=query) |
            Q(description__icontains=query)
        )
    if location:
        jobs = jobs.filter(location__icontains=location)
    if experience:
        try:
            exp_int = int(experience)
            jobs    = jobs.filter(min_experience__lte=exp_int, max_experience__gte=exp_int)
        except ValueError:
            pass

    return render(request, 'core/search_results.html', {
        'jobs':       jobs,
        'query':      query,
        'location':   location,
        'experience': experience,
        'total':      jobs.count(),
    })


# ===================== REGISTER =====================

def register_view(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            full_name     = form.cleaned_data["full_name"]
            email         = form.cleaned_data["email"]
            password      = form.cleaned_data["password"]
            mobile_number = form.cleaned_data["mobile_number"]
            work_status   = form.cleaned_data["work_status"]

            base_username = email.split("@")[0]
            username      = base_username
            counter       = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1

            user = User.objects.create_user(
                username=username, email=email,
                password=password, first_name=full_name,
            )
            UserProfile.objects.create(
                user=user, full_name=full_name,
                mobile_number=mobile_number, work_status=work_status,
            )
            messages.success(request, f"✅ Registration successful! Welcome {full_name}. Please login.")
            return redirect("login")
        else:
            messages.error(request, "❌ Registration failed. Please fix errors below.")
    else:
        form = RegisterForm()
    return render(request, "core/register.html", {"form": form})


# ===================== LOGIN =====================

def login_view(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            email    = form.cleaned_data["email"]
            password = form.cleaned_data["password"]
            try:
                user_obj = User.objects.get(email__iexact=email)
            except User.DoesNotExist:
                messages.error(request, "❌ No account found with this email.")
                return render(request, "core/login.html", {"form": form})

            user = authenticate(request, username=user_obj.username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"✅ Welcome back, {user.first_name}!")
                try:
                    if user.userprofile.role == 'employer':
                        return redirect('employer_dashboard')
                    else:
                        return redirect('index')
                except UserProfile.DoesNotExist:
                    return redirect('index')
            else:
                messages.error(request, "❌ Incorrect password.")
        else:
            messages.error(request, "❌ Please fill all fields correctly.")
    else:
        form = LoginForm()
    return render(request, "core/login.html", {"form": form})


# ===================== LOGOUT =====================

def logout_view(request):
    logout(request)
    messages.success(request, "✅ Logged out successfully.")
    return redirect("login")


def employer_login_page(request):
    return render(request, 'core/employer_login.html')


# ===================== JOB LISTING PAGES =====================
# ✅ Duplicates removed — one view per category kept
# Removed: it_jobs_page (duplicate of software_it_jobs_page)
#          marketingjobs_page (duplicate of marketing_jobs_page)
#          salesjobs_page (duplicate of sales_jobs_page)
#          banking_financejobs_page (near-duplicate of banking_finance_jobs_page)
#          fresherjobs_page (duplicate of fresher_jobs_page)
#          foreign_mnc_jobs_page (duplicate of mnc_jobs_page)
#          data_science_jobs_page (duplicate of datascience_jobs_page)
#          engineeringjobs_page (duplicate of engineering_jobs_page)
#          hr_jobs_page (duplicate of human_resources_jobs_page)

@login_required
def remote_jobs_page(request):
    return job_list_view(request, Job.objects.all().order_by('-id'), 'core/remote_jobs.html')

@login_required
def mnc_jobs_page(request):
    return job_list_view(request, Job.objects.filter(company_type__icontains='mnc').order_by('-id'), 'core/mnc_jobs.html')

@login_required
def banking_finance_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Banking').order_by('-id'), 'core/banking_finance_jobs.html')

@login_required
def startup_jobs_page(request):
    return job_list_view(request, Job.objects.filter(company_type__icontains='startup').order_by('-id'), 'core/startup_jobs.html')

@login_required
def software_it_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='IT').order_by('-id'), 'core/software_it_jobs.html')

@login_required
def internship_jobs_page(request):
    return job_list_view(request, Job.objects.filter(job_type='internship').order_by('-id'), 'core/internship_jobs.html')

@login_required
def engineering_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Engineering').order_by('-id'), 'core/engineering_jobs.html')

@login_required
def marketing_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Marketing').order_by('-id'), 'core/marketing_jobs.html')

@login_required
def fortune_jobs_page(request):
    return job_list_view(request, Job.objects.filter(company_type__icontains='fortune').order_by('-id'), 'core/fortune_jobs.html')

@login_required
def human_resources_jobs_page(request):
    return job_list_view(request, Job.objects.filter(Q(category__icontains='Human Resources') | Q(category__icontains='HR')).order_by('-id'), 'core/human_resources_jobs.html')

@login_required
def project_management_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Project Management').order_by('-id'), 'core/project_management_jobs.html')

@login_required
def finance_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Finance').order_by('-id'), 'core/finance_jobs.html')

@login_required
def operations_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Operations').order_by('-id'), 'core/operations_jobs.html')

@login_required
def supply_chain_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Supply Chain').order_by('-id'), 'core/supply_chain_jobs.html')

@login_required
def work_from_home_jobs_page(request):
    return job_list_view(request, Job.objects.filter(work_mode__icontains='Work From Home').order_by('-id'), 'core/work_from_home_jobs.html')

@login_required
def analytics_bi_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Analytics').order_by('-id'), 'core/analytics_bi_jobs.html')

@login_required
def datascience_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Data Science').order_by('-id'), 'core/datascience_jobs.html')

@login_required
def sales_jobs_page(request):
    return job_list_view(request, Job.objects.filter(category__icontains='Sales').order_by('-id'), 'core/sales_jobs.html')

@login_required
def fresher_jobs_page(request):
    return job_list_view(request, Job.objects.filter(min_experience=0).order_by('-id'), 'core/fresher_jobs.html')

@login_required
def walk_in_jobs_page(request):
    return job_list_view(request, Job.objects.filter(job_type__icontains='Walk In').order_by('-id'), 'core/walk_in_jobs.html')

@login_required
def part_time_jobs_page(request):
    return job_list_view(request, Job.objects.filter(job_type__icontains='Part Time').order_by('-id'), 'core/part_time_jobs.html')

@login_required
def delhi_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Delhi').order_by('-id'), 'core/delhi_jobs.html')

@login_required
def mumbai_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Mumbai').order_by('-id'), 'core/mumbai_jobs.html')

@login_required
def bangalore_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Bangalore').order_by('-id'), 'core/bangalore_jobs.html')

@login_required
def hyderabad_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Hyderabad').order_by('-id'), 'core/hyderabad_jobs.html')

@login_required
def chennai_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Chennai').order_by('-id'), 'core/chennai_jobs.html')

@login_required
def pune_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Pune').order_by('-id'), 'core/pune_jobs.html')

@login_required
def kolkata_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Kolkata').order_by('-id'), 'core/kolkata_jobs.html')

@login_required
def ahmedabad_jobs_page(request):
    return job_list_view(request, Job.objects.filter(location__icontains='Ahmedabad').order_by('-id'), 'core/ahmedabad_jobs.html')


# ===================== COMPANY PAGES =====================

@login_required
def company_unicorn(request):
    base_qs = Job.objects.filter(company_type__icontains='unicorn').order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Unicorn Companies Actively Hiring', 'page_category': 'Unicorns', 'clear_url': 'company_unicorn'})
    return render(request, 'core/company_unicorn.html', context)

@login_required
def company_mnc_jobs_page(request):
    base_qs = Job.objects.filter(Q(company_type__icontains='mnc') | Q(company_type__icontains='multinational')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'MNC Companies Actively Hiring', 'page_category': 'MNCs', 'clear_url': 'company_mnc'})
    return render(request, 'core/company_mnc_jobs.html', context)

@login_required
def company_startups_jobs_page(request):
    base_qs = Job.objects.filter(Q(company_type__icontains='startup') | Q(company_type__icontains='start-up')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Startup Companies Actively Hiring', 'page_category': 'Startups', 'clear_url': 'company_startups'})
    return render(request, 'core/company_startups_jobs.html', context)

@login_required
def company_product_based_jobs_page(request):
    base_qs = Job.objects.filter(Q(category__icontains='product') | Q(category__icontains='product based')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Product Based Companies Hiring', 'page_category': 'Product Based', 'clear_url': 'company_product_based'})
    return render(request, 'core/company_product_based_jobs.html', context)

@login_required
def company_internet_jobs_page(request):
    base_qs = Job.objects.filter(Q(category__icontains='internet') | Q(industry__icontains='internet')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Internet Companies Hiring', 'page_category': 'Internet', 'clear_url': 'company_internet'})
    return render(request, 'core/company_internet_jobs.html', context)

@login_required
def company_top_companies_jobs_page(request):
    base_qs = Job.objects.filter(Q(category__icontains='top company') | Q(company_type__icontains='mnc')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Top Companies Hiring', 'page_category': 'Top Companies', 'clear_url': 'company_top_companies'})
    return render(request, 'core/company_top_companies_jobs.html', context)

@login_required
def company_it_companies_jobs_page(request):
    base_qs = Job.objects.filter(Q(industry__icontains='it') | Q(category__icontains='it') | Q(company_type__icontains='it services')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'IT Companies Hiring', 'page_category': 'IT Companies', 'clear_url': 'company_it_companies'})
    return render(request, 'core/company_it_companies_jobs.html', context)

@login_required
def company_fintech_companies_jobs_page(request):
    base_qs = Job.objects.filter(Q(industry__icontains='fintech') | Q(category__icontains='fintech')).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Fintech Companies Hiring', 'page_category': 'Fintech', 'clear_url': 'company_fintech_companies'})
    return render(request, 'core/company_fintech_companies_jobs.html', context)

@login_required
def company_sponsored_companies_jobs_page(request):
    base_qs = Job.objects.filter(is_sponsored=True).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Sponsored Companies Hiring', 'page_category': 'Sponsored Companies', 'clear_url': 'company_sponsored_companies'})
    return render(request, 'core/company_sponsored_companies_jobs.html', context)

@login_required
def company_featured_companies_jobs_page(request):
    base_qs = Job.objects.filter(is_featured=True).order_by('-id')
    _, context = _build_company_filter_context(request, base_qs)
    context.update({'page_title': 'Featured Companies Hiring', 'page_category': 'Featured Companies', 'clear_url': 'company_featured_companies'})
    return render(request, 'core/company_featured_companies_jobs.html', context)


def company_jobs_page(request, company):
    jobs = Job.objects.filter(company__iexact=company).order_by('-created_at')
    jobs, selected = apply_all_filters(jobs, request)

    try:
        company_profile = CompanyProfile.objects.get(employer__username__iexact=company)
    except CompanyProfile.DoesNotExist:
        company_profile = None

    first_job        = Job.objects.filter(company__iexact=company).first()
    all_company_jobs = Job.objects.filter(company__iexact=company)
    counts           = get_filter_counts(all_company_jobs)

    return render(request, 'core/company_jobs.html', {
        'jobs':            jobs,
        'company':         company,
        'company_profile': company_profile,
        'first_job':       first_job,
        'total_jobs':      jobs.count(),
        **selected,
        **counts,
    })


# ===================== SAVE / APPLY / SAVED JOBS =====================

@candidate_required
def save_job(request, job_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    job   = get_object_or_404(Job, id=job_id)
    saved, created = SavedJob.objects.get_or_create(user=request.user, job=job)
    if created:
        return JsonResponse({'status': 'saved', 'message': 'Saved successfully!'})
    saved.delete()
    return JsonResponse({'status': 'removed', 'message': 'Removed from saved jobs!'})


@candidate_required
@transaction.atomic                          # ✅ all DB writes succeed or all rollback
def apply_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)

    if Application.objects.filter(applicant=request.user, job=job).exists():
        messages.warning(request, "You already applied for this job.")
        return redirect('applied_jobs')

    if request.method == 'POST':
        phone_number   = request.POST.get('phone_number', '')
        resume         = request.FILES.get('resume')
        skills         = request.POST.get('skills', '')
        location       = request.POST.get('location', '')
        experience_raw = request.POST.get('experience', '')

        if str(experience_raw).strip().lower() in ['fresher', 'freshers', '']:
            experience = 0
        else:
            try:
                experience = int(experience_raw)
            except (ValueError, TypeError):
                experience = 0

        Application.objects.create(
            applicant=request.user, job=job,
            phone_number=phone_number, resume=resume,
            status='Applied', skills=skills,
            location=location, experience=experience,
        )
        messages.success(request, "✅ Application submitted successfully!")
    return render(request, 'core/apply_job.html', {
            'job': job,
            'experience_range': range(0, 31),
        })


@candidate_required
def saved_jobs_page(request):
    saved = SavedJob.objects.filter(user=request.user).select_related('job').order_by('-saved_at')
    return render(request, 'core/saved_jobs.html', {'saved_jobs': saved})


def job_detail(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    if job.min_salary is not None and job.min_salary <= 0:
        job.min_salary = None
    if job.max_salary is not None and job.max_salary <= 0:
        job.max_salary = None
    return render(request, 'core/job_detail.html', {'job': job})


def remove_saved_job(request, saved_job_id):
    saved_job = get_object_or_404(SavedJob, id=saved_job_id, user=request.user)
    saved_job.delete()
    return redirect('saved_jobs')


@candidate_required
def applied_jobs_page(request):
    applied_jobs = Application.objects.filter(
        applicant=request.user
    ).select_related('job').order_by('-applied_at')
    return render(request, 'core/applied_jobs.html', {'applied_jobs': applied_jobs})


# ===================== EMPLOYER REGISTER / LOGIN =====================

def employer_register(request):
    if request.method == "POST":
        form = EmployerRegisterForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            email    = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user     = User.objects.create_user(username=username, email=email, password=password)
            profile  = form.save(commit=False)
            profile.user = user
            profile.role = "employer"
            profile.save()
            return redirect('employer_login_page')
    else:
        form = EmployerRegisterForm()
    return render(request, 'core/employer_register.html', {'form': form})


def employer_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user     = authenticate(request, username=username, password=password)
        if user is not None and user.userprofile.role == "employer":
            login(request, user)
            return redirect('employer_dashboard')
        # ✅ Fixed: was HttpResponse plain text — now proper message + redirect
        messages.error(request, "❌ Only employers can login here. Please check your credentials.")
        return render(request, 'core/employer_login_page.html')
    return render(request, 'core/employer_login_page.html')


# ===================== EMPLOYER DASHBOARD =====================

@employer_required
def employer_dashboard(request):
    # ✅ N+1 FIXED — annotate() replaces per-job DB query loop
    # BEFORE: for job in jobs: job.application_count = Application.objects.filter(job=job).count()
    # AFTER:  single JOIN query gives count for all jobs at once
    jobs = Job.objects.filter(employer=request.user).annotate(
        application_count=Count('applications')
    )
    all_applications = Application.objects.filter(job__employer=request.user)

    experience = request.GET.get('experience')
    location   = request.GET.get('location')
    skill      = request.GET.get('skill')
    status     = request.GET.get('status')

    applications = Application.objects.filter(
        job__employer=request.user
    ).select_related('applicant', 'job').order_by('-applied_at')

    if experience:
        applications = applications.filter(experience__gte=experience)
    if location:
        applications = applications.filter(location__icontains=location)
    if skill:
        applications = applications.filter(skills__icontains=skill)
    if status:
        applications = applications.filter(status=status)

    pipeline = {
        'applied':     all_applications.filter(status='Applied').count(),
        'screening':   all_applications.filter(status='Screening').count(),
        'shortlisted': all_applications.filter(status='Shortlisted').count(),
        'interview':   all_applications.filter(status='Interview').count(),
        'technical':   all_applications.filter(status='Technical').count(),
        'hr':          all_applications.filter(status='HR').count(),
        'offer':       all_applications.filter(status='Offer').count(),
    }

    default_experience = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '10', '12', '15']
    db_experience = list(
        Application.objects.filter(job__employer=request.user)
        .exclude(experience__isnull=True)
        .values_list('experience', flat=True).distinct()
    )

    default_locations = ['Ahmedabad', 'Bangalore', 'Chennai', 'Delhi', 'Gurgaon', 'Hyderabad', 'Kolkata', 'Mumbai', 'Noida', 'Pune']
    db_locations = list(
        Application.objects.filter(job__employer=request.user)
        .exclude(location__isnull=True).exclude(location='')
        .values_list('location', flat=True).distinct()
    )

    default_skills = ['AWS', 'CSS', 'Data Science', 'Django', 'Docker', 'Flutter', 'HTML', 'Java', 'JavaScript', 'Machine Learning', 'MongoDB', 'Node.js', 'Python', 'React', 'SQL']
    db_skills = []
    for skills in Application.objects.filter(job__employer=request.user).exclude(skills__isnull=True).values_list('skills', flat=True):
        if isinstance(skills, list):
            db_skills.extend(skills)
        elif isinstance(skills, str) and skills:
            db_skills.append(skills)

    context = {
        'jobs':                 jobs,
        'applications':         applications,
        'active_jobs':          jobs.filter(is_active=True).count(),
        'total_jobs':           jobs.count(),
        'total_applications':   all_applications.count(),
        'shortlisted_count':    applications.filter(status='Shortlisted').count(),
        'interviews_scheduled': applications.filter(status='Interview').count(),
        'pending_reviews':      applications.filter(status='Applied').count(),
        'total_views':          sum(job.views for job in jobs),
        'recent_applicants':    applications.order_by('-applied_at')[:5],
        'pipeline':             pipeline,
        'unread_messages':      Message.objects.filter(receiver=request.user, is_read=False).count(),
        'experience_options': sorted(set(int(e) for e in default_experience if str(e).isdigit()) | set(db_experience)),
        'location_options':     sorted(set(default_locations + db_locations)),
        'skill_options': sorted(set(default_skills) | set(db_skills)),
        'status_choices':       ['Applied', 'Screening', 'Shortlisted', 'Interview', 'Technical', 'HR', 'Offer', 'Rejected'],
        'selected_experience':  experience,
        'selected_location':    location,
        'selected_skill':       skill,
        'selected_status':      status,
    }
    return render(request, 'core/employer_dashboard.html', context)


@employer_required
def dashboard_realtime_data(request):
    jobs   = Job.objects.filter(employer=request.user)
    recent = Application.objects.filter(job__in=jobs).select_related('applicant', 'job').order_by('-applied_at')[:10]
    data   = {
        'total_applications': Application.objects.filter(job__in=jobs).count(),
        'applicants': [
            {
                'name':   app.applicant.get_full_name(),
                'job':    app.job.title,
                'date':   app.applied_at.strftime('%d %b %Y'),
                'status': app.status,
            }
            for app in recent
        ],
    }
    return JsonResponse(data)


@employer_required
def update_application_status(request, app_id, new_status):
    application = get_object_or_404(Application, id=app_id)
    if application.job.employer != request.user:
        messages.error(request, "Access denied.")
        return redirect('employer_dashboard')

    valid_statuses = ['Applied', 'Screening', 'Shortlisted', 'Interview', 'Technical', 'HR', 'Offer', 'Rejected']
    if new_status not in valid_statuses:
        messages.error(request, f"Invalid status: {new_status}")
        return redirect('employer_dashboard')

    application.status = new_status
    application.save()
    candidate_name = application.applicant.get_full_name() or application.applicant.username
    messages.success(request, f"✅ {candidate_name}'s status updated to '{new_status}'.")
    return redirect('employer_dashboard')


# ===================== POST / MANAGE JOBS =====================

@employer_required
def post_job(request):
    if request.method == 'POST':
        form = JobForm(request.POST, request.FILES)
        if form.is_valid():
            job          = form.save(commit=False)
            job.employer = request.user
            job.save()
            messages.success(request, '✅ Job posted successfully!')
            return redirect('post_job')
    else:
        form = JobForm()

    jobs = Job.objects.filter(employer=request.user).order_by('-created_at')
    return render(request, 'core/post_job.html', {'form': form, 'jobs': jobs})


@employer_required
def manage_jobs(request):
    jobs              = Job.objects.filter(employer=request.user).order_by('-created_at')
    recent_applicants = Application.objects.filter(
        job__employer=request.user
    ).select_related('applicant', 'job').order_by('-applied_at')[:5]

    context = {
        'jobs':               jobs,
        'total_jobs':         jobs.count(),
        'active_jobs':        jobs.filter(is_active=True, status='active').count(),
        'closed_jobs':        jobs.filter(status='closed').count(),
        'draft_jobs':         jobs.filter(status='draft').count(),
        'total_applications': Application.objects.filter(job__employer=request.user).count(),
        'total_views':        sum(job.views for job in jobs),
        'recent_applicants':  recent_applicants,
    }
    return render(request, 'core/manage_jobs.html', context)


@employer_required
def toggle_job_status(request, job_id):
    if request.method == 'POST':
        job           = get_object_or_404(Job, id=job_id, employer=request.user)
        job.is_active = not job.is_active
        job.status    = 'active' if job.is_active else 'closed'
        job.save()
        messages.success(request, f"✅ '{job.title}' is now {'Active' if job.is_active else 'Closed'}.")
    return redirect('manage_jobs')


@employer_required
def edit_job(request, job_id):
    job = get_object_or_404(Job, id=job_id, employer=request.user)
    if request.method == 'POST':
        form = JobForm(request.POST, request.FILES, instance=job)
        if form.is_valid():
            updated_job          = form.save(commit=False)
            updated_job.employer = request.user
            updated_job.save()
            messages.success(request, f'✅ "{job.title}" updated successfully!')
            return redirect('manage_jobs')
        messages.error(request, '❌ Please fix the errors below.')
    else:
        form = JobForm(instance=job)
    return render(request, 'core/edit_job.html', {'form': form, 'job': job})


@employer_required
def delete_job(request, job_id):
    job = get_object_or_404(Job, id=job_id, employer=request.user)
    job.delete()
    return redirect('employer_dashboard')


@employer_required
def view_applications(request, job_id):
    job          = get_object_or_404(Job, id=job_id)
    applications = Application.objects.filter(
        job=job
    ).select_related('applicant', 'job').order_by('-applied_at')
    return render(request, 'core/view_applications.html', {'job': job, 'applications': applications})


# ===================== APPLICANTS =====================

@employer_required
def applicants(request):
    applications = Application.objects.filter(
        job__employer=request.user
    ).select_related('applicant', 'job').order_by('-applied_at')

    apps_list = [{
        'id':         a.id,
        'name':       a.applicant.get_full_name() or a.applicant.username,
        'email':      a.applicant.email,
        'phone':      a.phone_number or '',
        'job_title':  a.job.title,
        'job_id':     a.job.id,
        'status':     a.status,
        'skills':     a.skills or '',
        'location':   a.location or '',
        'experience': a.experience or '',
        'resume':     a.resume.url if a.resume else '',
        'applied_at': a.applied_at.strftime('%d %b %Y'),
    } for a in applications]

    employer_jobs = Job.objects.filter(employer=request.user).order_by('title')

    context = {
        'applications_json':  apps_list,
        'total_applications': applications.count(),
        'pending_count':      applications.filter(status='Applied').count(),
        'shortlisted_count':  applications.filter(status='Shortlisted').count(),
        'interview_count':    applications.filter(status='Interview').count(),
        'rejected_count':     applications.filter(status='Rejected').count(),
        'employer_jobs':      employer_jobs,
        'total_jobs':         employer_jobs.count(),
        'status_choices':     ['Applied', 'Screening', 'Shortlisted', 'Interview', 'Technical', 'HR', 'Offer', 'Rejected'],
    }
    return render(request, 'core/applicants.html', context)


@employer_required
def update_status(request, app_id, status):
    application = get_object_or_404(Application, id=app_id)
    if application.job.employer != request.user:
        messages.error(request, "❌ Access denied.")
        return redirect('applicants')

    if request.method == 'POST':
        status         = request.POST.get('new_status')
        valid_statuses = ['Applied', 'Screening', 'Shortlisted', 'Interview', 'Technical', 'HR', 'Offer', 'Rejected']
        if not status or status not in valid_statuses:
            messages.error(request, f"❌ Invalid status: {status}")
            return redirect('applicants')
        application.status = status
        application.save()
        candidate_name = application.applicant.get_full_name() or application.applicant.username
        messages.success(request, f"✅ {candidate_name} marked as {status}.")
    return redirect('applicants')


@employer_required
def shortlist_candidate(request, app_id):
    application = get_object_or_404(Application, id=app_id)
    if application.job.employer != request.user:
        messages.error(request, "Access denied.")
        return redirect('employer_dashboard')
    application.status = "Shortlisted"
    application.save()
    messages.success(request, "✅ Candidate shortlisted successfully.")
    return redirect('employer_dashboard')


@employer_required
def shortlisted_candidates(request):
    shortlisted = Application.objects.filter(
        job__employer=request.user, status='Shortlisted'
    ).select_related('applicant', 'job').order_by('-applied_at')
    
    for app in shortlisted:
        if not app.skills:
            app.skills_list = []
        elif isinstance(app.skills, list):
            app.skills_list = app.skills
        else:
            app.skills_list = app.skills.split(',')
    
    return render(request, 'core/shortlisted.html', {'shortlisted': shortlisted})


@employer_required
def reject_candidate(request, app_id):
    application = get_object_or_404(Application, id=app_id)
    if application.job.employer != request.user:
        messages.error(request, "Access denied.")
        return redirect('employer_dashboard')
    application.status = "Rejected"
    application.save()
    messages.success(request, "Candidate rejected.")
    return redirect('employer_dashboard')


@employer_required
def recruiter_applications(request):
    applications = ApplyJob.objects.select_related('user', 'job').all().order_by('-applied_at')
    return render(request, 'core/recruiter_applications.html', {'applications': applications})


# ===================== INTERVIEWS =====================

@employer_required
def interviews(request):
    today          = timezone.now().date()
    interview_list = Interview.objects.filter(
        job__employer=request.user
    ).select_related('candidate', 'job').order_by('interview_date', 'interview_time')

    return render(request, 'core/interviews.html', {
        'interviews':      interview_list,
        'upcoming_count':  interview_list.filter(interview_date__gte=today, status='Scheduled').count(),
        'completed_count': interview_list.filter(status='Completed').count(),
        'cancelled_count': interview_list.filter(status='Cancelled').count(),
        'round_types':     interview_list.values_list('round_type', flat=True).distinct(),
        'statuses':        interview_list.values_list('status', flat=True).distinct(),
    })


@employer_required
def schedule_interview(request, app_id):
    application = get_object_or_404(Application, id=app_id)
    if application.job.employer != request.user:
        messages.error(request, "Access denied.")
        return redirect('shortlisted_candidates')

    if request.method == 'POST':
        Interview.objects.create(
            candidate      = application.applicant,
            job            = application.job,
            round_type     = request.POST.get('round_type', 'Technical'),
            interview_date = request.POST.get('interview_date'),
            interview_time = request.POST.get('interview_time'),
            meeting_link   = request.POST.get('meeting_link', ''),
            status         = 'Scheduled',
        )
        application.status = 'Interview Scheduled'
        application.save()
        messages.success(request, f"✅ Interview scheduled for {application.applicant.get_full_name() or application.applicant.username}!")
        return redirect('interviews')

    return render(request, 'core/schedule_interview.html', {'application': application})


# ===================== MESSAGES =====================

@employer_required
def inbox_messages(request):
    all_messages = Message.objects.filter(
        Q(sender=request.user) | Q(receiver=request.user)
    ).select_related('sender', 'receiver').order_by('created_at')

    conversations = {}
    for msg in all_messages:
        other = msg.receiver if msg.sender == request.user else msg.sender
        if other.id not in conversations:
            conversations[other.id] = {
                'candidate':         other,
                'messages':          [],
                'unread_count':      0,
                'last_message':      '',
                'last_message_time': msg.created_at,
            }
        conversations[other.id]['messages'].append(msg)
        conversations[other.id]['last_message']      = msg.message
        conversations[other.id]['last_message_time'] = msg.created_at
        if not msg.is_read and msg.receiver == request.user:
            conversations[other.id]['unread_count'] += 1

    return render(request, 'core/messages.html', {
        'conversations': list(conversations.values()),
        'unread_count':  Message.objects.filter(receiver=request.user, is_read=False).count(),
    })


@employer_required
def send_message(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data         = json.loads(request.body)
        receiver_id  = data.get('receiver_id')
        message_text = data.get('message', '').strip()
        if not receiver_id or not message_text:
            return JsonResponse({'success': False, 'error': 'Missing receiver or message'})
        receiver = User.objects.get(id=receiver_id)
        msg      = Message.objects.create(
            sender=request.user, receiver=receiver,
            message=message_text, is_read=False,
        )
        return JsonResponse({
            'success':   True,
            'message_id': msg.id,
            'text':       msg.message,
            'time':       msg.created_at.strftime('%H:%M'),
            'sender_id':  request.user.id,
        })
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Candidate not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@employer_required
def fetch_messages(request, candidate_id):
    messages_qs = Message.objects.filter(
        Q(sender=request.user, receiver_id=candidate_id) |
        Q(sender_id=candidate_id, receiver=request.user)
    ).order_by('created_at')
    messages_qs.filter(receiver=request.user, is_read=False).update(is_read=True)
    data = [
        {'id': m.id, 'text': m.message, 'sender_id': m.sender.id, 'time': m.created_at.strftime('%H:%M')}
        for m in messages_qs
    ]
    return JsonResponse({'success': True, 'messages': data})


# ===================== REPORTS =====================

@employer_required
def reports(request):
    """
    ✅ Delegates to ReportService — no raw queries here.
    """
    date_from    = request.GET.get('date_from', '')
    date_to      = request.GET.get('date_to', '')
    selected_job = request.GET.get('job_title', '')

    context = ReportService.get_dashboard_context(
        user=request.user,
        date_from=date_from,
        date_to=date_to,
        selected_job=selected_job,
    )
    return render(request, 'core/reports.html', context)


@employer_required
def export_reports_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    # ✅ ReportService used — no duplicate queries
    summary      = ReportService.get_summary(request.user)
    applications = ReportService.get_all_applications(request.user)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story  = [Paragraph("HireHub — Recruitment Report", styles['Title']), Spacer(1, 12)]

    summary_data = [
        ['Metric', 'Count'],
        ['Total Jobs Posted',  str(summary['total_jobs'])],
        ['Total Applications', str(summary['total_applications'])],
        ['Shortlisted',        str(summary['shortlisted_count'])],
        ['Interviews',         str(summary['interview_count'])],
        ['Rejected',           str(summary['rejected_count'])],
    ]
    t = Table(summary_data, colWidths=[300, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (0, 0), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING',    (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 24))
    story.append(Paragraph("All Applications", styles['Heading2']))
    story.append(Spacer(1, 8))

    app_data = [['Candidate', 'Job Title', 'Status', 'Applied Date']]
    for app in applications:
        app_data.append([
            app.applicant.get_full_name() or app.applicant.username,
            app.job.title,
            app.status,
            app.applied_at.strftime('%d %b %Y'),
        ])

    at = Table(app_data, colWidths=[150, 150, 100, 100])
    at.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    story.append(at)
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="hirehub_report.pdf"'
    return response


@employer_required
def export_reports_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io

    # ✅ ReportService used — no duplicate queries
    summary   = ReportService.get_summary(request.user)
    apps      = ReportService.get_all_applications(request.user)
    top_jobs  = ReportService.get_top_jobs_for_excel(request.user)

    wb          = openpyxl.Workbook()
    header_fill = PatternFill(start_color='4f46e5', end_color='4f46e5', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    ws1 = wb.active
    ws1.title = 'Summary'
    ws1.append(['Metric', 'Count'])
    ws1.append(['Total Jobs Posted',  summary['total_jobs']])
    ws1.append(['Total Applications', summary['total_applications']])
    ws1.append(['Shortlisted',        summary['shortlisted_count']])
    ws1.append(['Interviews',         summary['interview_count']])
    ws1.append(['Rejected',           summary['rejected_count']])
    style_header(ws1)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 15

    ws2 = wb.create_sheet('Applications')
    ws2.append(['Candidate', 'Email', 'Job Title', 'Status', 'Location', 'Experience', 'Applied Date'])
    style_header(ws2)
    for app in apps:
        ws2.append([
            app.applicant.get_full_name() or app.applicant.username,
            app.applicant.email,
            app.job.title,
            app.status,
            app.location or '—',
            app.experience or '—',
            app.applied_at.strftime('%d %b %Y'),
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 20

    ws3 = wb.create_sheet('Top Jobs')
    ws3.append(['Job Title', 'Applications', 'Shortlisted', 'Views', 'Hire Rate %'])
    style_header(ws3)
    for job in top_jobs:
        ws3.append([job['title'], job['app_count'], job['sc'], job['views'], job['hire_rate']])
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws3.column_dimensions[col].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="hirehub_report.xlsx"'
    return response


# ===================== COMPANY PROFILE & SETTINGS =====================

@employer_required
def company_profile(request):
    """
    ✅ Uses CompanyProfileForm — fat view logic removed.
    Form handles all field assignment and validation cleanly.
    """
    profile, _ = CompanyProfile.objects.get_or_create(employer=request.user)

    if request.method == 'POST':
        form = CompanyProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Company profile saved successfully!')
            return redirect('company_profile')
        messages.error(request, '❌ Please fix the errors below.')
    else:
        form = CompanyProfileForm(instance=profile)

    return render(request, 'core/company_profile.html', {'form': form, 'profile': profile})


@employer_required
def settings(request):
    """
    ✅ Delegates to SettingsService — view no longer touches 3 models directly.
    """
    employer_settings, _ = EmployerSettings.objects.get_or_create(employer=request.user)

    if request.method == 'POST':
        SettingsService.update_all(request.user, request.POST, request.FILES)
        messages.success(request, '✅ Settings saved successfully!')
        return redirect('settings')

    return render(request, 'core/settings.html', {'employer_settings': employer_settings})


@employer_required
def change_password(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    from django.contrib.auth import update_session_auth_hash
    try:
        data             = json.loads(request.body)
        current_password = data.get('current_password', '')
        new_password     = data.get('new_password', '')
        if not request.user.check_password(current_password):
            return JsonResponse({'success': False, 'error': 'Current password is incorrect.'})
        if len(new_password) < 6:
            return JsonResponse({'success': False, 'error': 'New password must be at least 6 characters.'})
        request.user.set_password(new_password)
        request.user.save()
        update_session_auth_hash(request, request.user)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@employer_required
def deactivate_account(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        Job.objects.filter(employer=request.user).update(is_active=False)
        request.user.is_active = False
        request.user.save()
        logout(request)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ===================== MISC =====================

@login_required
def all_jobs(request):
    return render(request, 'core/all_jobs.html', {'jobs': Job.objects.all()})


def about(request):             return render(request, 'core/about.html')
def careers(request):           return render(request, 'core/careers.html')
def employer_home(request):     return render(request, 'core/employer_home.html')
def sitemap(request):           return render(request, 'core/sitemap.html')
def credits(request):           return render(request, 'core/credits.html')
def help_center(request):       return render(request, 'core/help_center.html')
def summons_notices(request):   return render(request, 'core/summons_notices.html')
def grievances(request):        return render(request, 'core/grievances.html')
def report_issue(request):      return render(request, 'core/report_issue.html')
def privacy_policy(request):    return render(request, 'core/privacy_policy.html')
def terms_conditions(request):  return render(request, 'core/terms_conditions.html')
def fraud_alert(request):       return render(request, 'core/fraud_alert.html')
def trust_safety(request):      return render(request, 'core/trust_safety.html')
def search_jobs(request):       return render(request, 'core/search_jobs.html')
def browser_companies(request): return render(request, 'core/browser_companies.html')
def resume_builder(request):    return render(request, 'core/resume_builder.html')
def career_advice(request):     return render(request, 'core/career_advice.html')
def salary_calculator(request): return render(request, 'core/salary_calculator.html')
def hiring_solutions(request):  return render(request, 'core/hiring_solutions.html')
def view_plans(request):        return render(request, 'core/view_plans.html')